import sys
import json
import pdfplumber
import pandas as pd
import os
from openpyxl import load_workbook

def extract_and_update(pdf_path, excel_path):
    extraction_results = {
        "rep_code": None,
        "fee": None,
        "status": "pending"
    }

    try:
        # 1. Extract Data
        with pdfplumber.open(pdf_path) as pdf:
            # Page 1: Rep Code
            if len(pdf.pages) > 0:
                p1_text = pdf.pages[0].extract_text()
                # Simple parsing strategy: Look for "Rep Code" and take next word/token
                # Adjust regex or logic as needed based on real PDF format
                if "Rep Code" in p1_text:
                    # Find the line containing "Rep Code"
                    for line in p1_text.split('\n'):
                        if "Rep Code" in line:
                            # Assuming format "Rep Code: XYZ" or "Rep Code XYZ"
                            parts = line.split("Rep Code")
                            if len(parts) > 1:
                                extraction_results["rep_code"] = parts[1].strip(": ").split()[0] # Take first word after label
                                break
            
            # Page 14: Fee (Index 13)
            if len(pdf.pages) >= 14:
                p14_text = pdf.pages[13].extract_text()
                if "Fee" in p14_text:
                     for line in p14_text.split('\n'):
                        if "Fee" in line:
                            parts = line.split("Fee")
                            if len(parts) > 1:
                                extraction_results["fee"] = parts[1].strip(": ").split()[0]
                                break
            else:
                # If less than 14 pages, check last page or log warning? 
                # For now strict adherence to prompt "page 14"
                pass

        # 2. Update Excel
        if not os.path.exists(excel_path):
            return {"error": f"Excel file not found at {excel_path}"}

        # Using openpyxl to modify existing file without overwriting entire structure if possible, 
        # but pandas is easier for tabular data. 
        # Prompt: "cell a1 says "rep code" cell a2 says "fee"" 
        # This implies A1/A2 are headers/labels. Values should likely go in B1/B2.
        
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Verify headers (Optional, but good safety)
        # if ws['A1'].value != "rep code": ...
        
        # Write values
        if extraction_results["rep_code"]:
            ws['B1'] = extraction_results["rep_code"]
        
        if extraction_results["fee"]:
            ws['B2'] = extraction_results["fee"]
            
        wb.save(excel_path)
        
        extraction_results["status"] = "success"
        return extraction_results

    except Exception as e:
        return {"error": str(e), "status": "failed"}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Missing arguments"}))
        sys.exit(1)
        
    pdf_path = sys.argv[1]
    # Hardcoded or passed arg? User gave specific path.
    excel_path = r"C:\Users\JoshuaBajorek\Desktop\Claude Code\Claude Code.xlsx"
    
    result = extract_and_update(pdf_path, excel_path)
    print(json.dumps(result))
